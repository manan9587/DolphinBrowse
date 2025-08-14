from __future__ import annotations

from pathlib import Path
from typing import Any, Dict

from openpyxl import Workbook


def analyze_pdf(file_path: str) -> Dict[str, Any]:
    try:
        from pypdf import PdfReader  # type: ignore
        reader = PdfReader(file_path)
        text = "".join(page.extract_text() or "" for page in reader.pages)
    except Exception:
        text = ""
    return {"text": text}


def analyze_docx(file_path: str) -> Dict[str, Any]:
    try:
        from docx import Document  # type: ignore
        doc = Document(file_path)
        text = "\n".join(p.text for p in doc.paragraphs)
    except Exception:
        text = ""
    return {"text": text}


def analyze_excel(file_path: str) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(file_path, read_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                lines.append(",".join(str(cell) for cell in row if cell is not None))
        text = "\n".join(lines)
    except Exception:
        text = ""
    return {"text": text}


class FileProcessor:
    def analyze(self, file_path: str) -> Dict[str, Any]:
        ext = Path(file_path).suffix.lower()
        if ext == ".pdf":
            return analyze_pdf(file_path)
        if ext == ".docx":
            return analyze_docx(file_path)
        if ext in {".xlsx", ".csv"}:
            return analyze_excel(file_path)
        return {"text": ""}

    def generate_remarks_excel(self, analysis: Dict[str, Any], output_path: str) -> str:
        text = analysis.get("text", "")
        wb = Workbook()
        ws = wb.active
        ws.title = "Remarks"
        ws.append(["Summary"])
        ws.append([f"Characters: {len(text)}"])
        wb.save(output_path)
        return output_path
